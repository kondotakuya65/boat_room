#!/usr/bin/env python3

from datetime import date, datetime
from typing import List, Dict, Tuple, Set
import re
import os
import tempfile
import requests
from openpyxl import load_workbook

from .client import get_gspread_client, get_sheets_service
from .color_dump import get_worksheet_colors


def _is_white(color: dict) -> bool:
    if not color:
        return False
    return color.get("r", 0) == 1 and color.get("g", 0) == 1 and color.get("b", 0) == 1


def _is_white_excel(fill) -> bool:
    """Check if Excel cell fill is white/empty"""
    if not fill or not fill.start_color:
        return True  # Default is white
    
    rgb = fill.start_color.rgb
    if not rgb or rgb == "00000000":  # No fill
        return True
    
    # Excel RGB is in ARGB format (8 characters)
    if len(rgb) == 8:
        # Remove alpha channel (first 2 characters)
        rgb = rgb[2:]
    
    if len(rgb) == 6:
        r, g, b = int(rgb[0:2], 16), int(rgb[2:4], 16), int(rgb[4:6], 16)
        # Check if it's white (RGB: 255, 255, 255)
        return r == 255 and g == 255 and b == 255
    
    return True


def _download_excel_file(sheet_link: str) -> str:
    """Download Excel file from Google Drive and return temporary file path"""
    # Extract file ID from Google Sheets URL
    file_id = sheet_link.split('/d/')[1].split('/')[0]
    
    # Download as Excel format
    download_url = f"https://docs.google.com/spreadsheets/d/{file_id}/export?format=xlsx"
    
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_path = temp_file.name
    temp_file.close()
    
    try:
        # Download the file
        response = requests.get(download_url)
        response.raise_for_status()
        
        # Write to temporary file
        with open(temp_path, 'wb') as f:
            f.write(response.content)
        
        return temp_path
    except Exception as e:
        # Clean up on error
        if os.path.exists(temp_path):
            os.unlink(temp_path)
        raise e


def _parse_excel_file(temp_path: str, boat_name: str) -> List[Dict]:
    """Parse Excel file and return room availability data"""
    workbook = load_workbook(temp_path, data_only=True)
    
    results: List[Dict] = []
    current_year = 2025
    
    # Build keyword->canonical mapping from config
    from ..config import BOAT_CATALOG as _CAT
    cfg_rooms = list((_CAT[boat_name].get("rooms") or {}).keys())
    keyword_to_canonical: Dict[str, str] = {}
    for name in cfg_rooms:
        keyword_to_canonical[name.upper().split()[0]] = name  # WAKATOBI -> Wakatobi
    
    # Find worksheets that look like years
    target_sheets = []
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().isdigit() and len(sheet_name.strip()) == 4:
            target_sheets.append(sheet_name)
    if not target_sheets:
        target_sheets = [workbook.sheetnames[0]]
    
    for worksheet_name in target_sheets:
        worksheet = workbook[worksheet_name]
        
        # Convert to list of lists for compatibility with existing logic
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        
        header_row_idx = _find_first_header_row(rows)
        if header_row_idx is None:
            continue
        range_row_idx = header_row_idx + 1
        header_vals = rows[header_row_idx]
        range_vals = rows[range_row_idx] if range_row_idx < len(rows) else []
        
        month_spans = _collect_month_spans(header_vals)
        if not month_spans:
            continue
        
        # Scan rows 10–25 in 3-row blocks; only start a block on known room label rows
        start_row = max(range_row_idx + 1, 10)
        end_row = min(len(rows), 26)
        
        i = start_row
        while i < end_row - 2:
            label = (rows[i][1] if len(rows[i]) > 1 else "").strip()
            if label:
                up = label.upper()
                matched_keyword = None
                for kw in keyword_to_canonical.keys():
                    if kw in up:
                        matched_keyword = kw
                        break
                if not matched_keyword:
                    i += 1  # Skip description/price or unrelated row
                    continue
                
                canonical_name = keyword_to_canonical[matched_keyword]
                
                # Three-row block
                row_block = [i, i + 1, i + 2]
                available_dates: List[date] = []
                
                # For each month span, scan all date-range tokens within its columns
                for span in month_spans:
                    mon = span["month"]
                    start_c = span["start_col"]
                    end_c = span["end_col"]
                    for col_idx in range(start_c, min(end_c + 1, len(range_vals))):
                        token = (range_vals[col_idx] or "").strip()
                        if not token or '-' not in token:
                            continue
                        parsed = _parse_date_range(token, mon, current_year)
                        if not parsed:
                            continue
                        start_str, _ = parsed
                        start_dt = datetime.strptime(start_str, "%Y/%m/%d").date()
                        
                        # Check white across 3 rows (all must be white to be available)
                        is_available = True
                        for r in row_block:
                            if r >= len(rows) or col_idx >= len(rows[r]):
                                is_available = False
                                break
                            # Get cell from Excel worksheet
                            cell = worksheet.cell(row=r+1, column=col_idx+1)  # Excel is 1-indexed
                            if not _is_white_excel(cell.fill):
                                is_available = False
                                break
                        
                        if is_available:
                            available_dates.append(start_dt)
                
                # Get room link from config
                from ..config import get_room_link
                room_link = get_room_link(boat_name, canonical_name)
                
                results.append({
                    "boat_name": boat_name,
                    "room_name": canonical_name,
                    "occupied": [],
                    "available_dates": available_dates,
                    "room_link": room_link,
                })
                
                i += 3
            else:
                i += 1
    
    return results


def _get_excel_all_sheet_start_dates(temp_path: str) -> Set[date]:
    """Get all start dates from Excel file"""
    workbook = load_workbook(temp_path, data_only=True)
    start_dates: Set[date] = set()
    current_year = 2025
    
    # Find worksheets that look like years
    target_sheets = []
    for sheet_name in workbook.sheetnames:
        if sheet_name.strip().isdigit() and len(sheet_name.strip()) == 4:
            target_sheets.append(sheet_name)
    if not target_sheets:
        target_sheets = [workbook.sheetnames[0]]
    
    for worksheet_name in target_sheets:
        worksheet = workbook[worksheet_name]
        
        # Convert to list of lists
        rows = []
        for row in worksheet.iter_rows(values_only=True):
            rows.append([str(cell) if cell is not None else "" for cell in row])
        
        header_row_idx = _find_first_header_row(rows)
        if header_row_idx is None:
            continue
        range_row_idx = header_row_idx + 1
        header_vals = rows[header_row_idx]
        range_vals = rows[range_row_idx] if range_row_idx < len(rows) else []
        month_spans = _collect_month_spans(header_vals)
        if not month_spans:
            continue
        for span in month_spans:
            mon = span["month"]
            for col_idx in range(span["start_col"], min(span["end_col"] + 1, len(range_vals))):
                token = (range_vals[col_idx] or "").strip()
                if not token or '-' not in token:
                    continue
                parsed = _parse_date_range(token, mon, current_year)
                if not parsed:
                    continue
                start_str, _ = parsed
                start_dates.add(datetime.strptime(start_str, "%Y/%m/%d").date())
    
    return start_dates


def _parse_date_range(token: str, month: int, year: int) -> Tuple[str, str] | None:
    token = (token or "").strip()
    if not token:
        return None
    if "-" not in token:
        return None
    parts = token.replace(" ", "").split("-")
    if len(parts) != 2:
        return None
    try:
        start_day = int(parts[0])
        end_day = int(parts[1])
    except ValueError:
        return None

    start_month = month
    end_month = month
    end_year = year

    # Cross-month: e.g., 30-1 means end in next month
    if end_day < start_day:
        if month == 12:
            end_month = 1
            end_year = year + 1
        else:
            end_month = month + 1

    start_str = f"{year:04d}/{start_month:02d}/{start_day:02d}"
    end_str = f"{end_year:04d}/{end_month:02d}/{end_day:02d}"
    return start_str, end_str


_MONTH_MAP = {
    "JAN": 1, "JANUARY": 1,
    "FEB": 2, "FEBRUARY": 2,
    "MAR": 3, "MARCH": 3,
    "APR": 4, "APRIL": 4,
    "MAY": 5,
    "JUN": 6, "JUNE": 6,
    "JUL": 7, "JULY": 7,
    "AUG": 8, "AUGUST": 8,
    "SEP": 9, "SEPT": 9, "SEPTEMBER": 9,
    "OCT": 10, "OCTOBER": 10,
    "NOV": 11, "NOVEMBER": 11,
    "DEC": 12, "DECEMBER": 12,
}


def _word_in_token(word: str, token: str) -> bool:
    return re.search(rf"\b{re.escape(word)}\b", token) is not None


def _find_first_header_row(rows: List[List[str]]) -> int | None:
    for i, row in enumerate(rows):
        upper = [str(c or "").strip().upper() for c in row]
        found = set()
        for cell in upper:
            for key in _MONTH_MAP.keys():
                if _word_in_token(key, cell):
                    found.add(_MONTH_MAP[key])
        if len(found) >= 2:
            return i
    return None


def _collect_month_spans(header_row_vals: List[str], merged_ranges: List[Dict] = None) -> List[Dict]:
    """From a single header row, find each month header column and define its span until next header.
    Uses merged cell information from Google Sheets API v4 to detect proper spans.
    Returns list of {month, start_col, end_col} sorted by start_col.
    """
    month_headers: List[Tuple[int, int]] = []  # (col_idx, month)
    for j, cell in enumerate(header_row_vals):
        token = (cell or "").strip().upper()
        for key, mon in _MONTH_MAP.items():
            if _word_in_token(key, token):
                month_headers.append((j, mon))
                break
    month_headers.sort(key=lambda x: x[0])
    spans: List[Dict] = []
    
    for idx, (col, mon) in enumerate(month_headers):
        # Default end column (next month header or end of row)
        next_month_col = month_headers[idx + 1][0] if idx + 1 < len(month_headers) else len(header_row_vals)
        end_col = next_month_col - 1
        
        # Check if this month header is part of a merged range
        if merged_ranges:
            for merge in merged_ranges:
                merge_start_row = merge.get('startRowIndex', 0)
                merge_end_row = merge.get('endRowIndex', 0)
                merge_start_col = merge.get('startColumnIndex', 0)
                merge_end_col = merge.get('endColumnIndex', 0)
                
                # Check if the month header column is within this merged range
                if (merge_start_row <= 8 <= merge_end_row and  # Header row is typically row 8 (0-indexed)
                    merge_start_col <= col < merge_end_col):
                    # This month header is part of a merged range
                    end_col = merge_end_col - 1
                    break
        
        spans.append({"month": mon, "start_col": col, "end_col": end_col})
    
    return spans


def parse_barakati_from_sheets(boat_name: str) -> List[Dict]:
    from ..config import BOAT_CATALOG

    if boat_name not in BOAT_CATALOG:
        return []

    sheet_link = BOAT_CATALOG[boat_name].get("sheet_link")
    if not sheet_link:
        return []

    # Check if this is an Excel file by trying to access it with gspread first
    try:
        gc = get_gspread_client()
        sheet = gc.open_by_url(sheet_link)
        # If we get here, it's a native Google Sheets document
        return _parse_google_sheets(sheet, boat_name)
    except Exception as e:
        # If gspread fails, try downloading as Excel file
        if "not supported for this document" in str(e):
            temp_path = None
            try:
                temp_path = _download_excel_file(sheet_link)
                return _parse_excel_file(temp_path, boat_name)
            finally:
                # Clean up temporary file
                if temp_path and os.path.exists(temp_path):
                    os.unlink(temp_path)
        else:
            raise e


def _parse_google_sheets(sheet, boat_name: str) -> List[Dict]:
    results: List[Dict] = []
    current_year = 2025

    # Iterate worksheets that look like years
    target_sheets = []
    for ws in sheet.worksheets():
        title = (ws.title or "").strip()
        if title.isdigit() and len(title) == 4:
            target_sheets.append(title)
    if not target_sheets:
        target_sheets = [sheet.worksheets()[0].title]

    debug_all_available: Dict[str, List[date]] = {}
    debug_by_date: Dict[str, List[str]] = {}

    # Build keyword->canonical mapping from config
    from ..config import BOAT_CATALOG as _CAT
    cfg_rooms = list((_CAT[boat_name].get("rooms") or {}).keys())
    keyword_to_canonical: Dict[str, str] = {}
    for name in cfg_rooms:
        keyword_to_canonical[name.upper().split()[0]] = name  # WAKATOBI -> Wakatobi

    for worksheet_title in target_sheets:
        rows = sheet.worksheet(worksheet_title).get('A1:ZZ1000')
        service = get_sheets_service()
        colors = get_worksheet_colors(service, sheet.id, worksheet_title)

        header_row_idx = _find_first_header_row(rows)
        if header_row_idx is None:
            continue
        range_row_idx = header_row_idx + 1
        header_vals = rows[header_row_idx]
        range_vals = rows[range_row_idx] if range_row_idx < len(rows) else []

        # Get merged cell information from Google Sheets API v4
        merged_ranges = []
        try:
            result = service.spreadsheets().get(
                spreadsheetId=sheet.id,
                ranges=[f'{worksheet_title}!A1:ZZ1000'],
                includeGridData=False
            ).execute()
            
            if 'sheets' in result and len(result['sheets']) > 0:
                sheet_data = result['sheets'][0]
                if 'merges' in sheet_data:
                    merged_ranges = sheet_data['merges']
        except Exception as e:
            # Fallback to heuristic method if API call fails
            pass

        month_spans = _collect_month_spans(header_vals, merged_ranges)
        if not month_spans:
            continue
        

        # Scan rows 10–25 in 3-row blocks; only start a block on known room label rows
        start_row = max(range_row_idx + 1, 10)
        end_row = min(len(rows), 26)

        i = start_row
        while i < end_row - 2:
            label = (rows[i][1] if len(rows[i]) > 1 else "").strip()
            if label:
                up = label.upper()
                matched_keyword = None
                for kw in keyword_to_canonical.keys():
                    if kw in up:
                        matched_keyword = kw
                        break
                if not matched_keyword:
                    i += 1  # Skip description/price or unrelated row
                    continue

                canonical_name = keyword_to_canonical[matched_keyword]

                # Three-row block
                row_block = [i, i + 1, i + 2]
                available_dates: List[date] = []

                # For each month span, scan all date-range tokens within its columns
                for span in month_spans:
                    mon = span["month"]
                    start_c = span["start_col"]
                    end_c = span["end_col"]
                    for col_idx in range(start_c, min(end_c + 1, len(range_vals))):
                        token = (range_vals[col_idx] or "").strip()
                        if not token or '-' not in token:
                            continue
                        parsed = _parse_date_range(token, mon, current_year)
                        if not parsed:
                            continue
                        start_str, _ = parsed
                        start_dt = datetime.strptime(start_str, "%Y/%m/%d").date()

                        # Check white across 3 rows (all must be white to be available)
                        is_available = True
                        for r in row_block:
                            if r >= len(colors) or col_idx >= len(colors[r]) or not _is_white(colors[r][col_idx]):
                                is_available = False
                                break
                        status = "AVAILABLE" if is_available else "OCCUPIED"
                        debug_by_date.setdefault(start_str, []).append(f"{canonical_name}={status}")
                        if is_available:
                            available_dates.append(start_dt)

                debug_all_available.setdefault(canonical_name, []).extend(available_dates)

                # Get room link from config
                from ..config import get_room_link
                room_link = get_room_link(boat_name, canonical_name)

                results.append({
                    "boat_name": boat_name,
                    "room_name": canonical_name,
                    "occupied": [],
                    "available_dates": available_dates,
                    "room_link": room_link,
                })

                i += 3
            else:
                i += 1

    return results


def get_barakati_all_sheet_start_dates(boat_name: str) -> Set[date]:
    from ..config import BOAT_CATALOG

    if boat_name not in BOAT_CATALOG:
        return set()

    sheet_link = BOAT_CATALOG[boat_name].get("sheet_link")
    if not sheet_link:
        return set()

    # Check if this is an Excel file by trying to access it with gspread first
    try:
        gc = get_gspread_client()
        sheet = gc.open_by_url(sheet_link)
        # If we get here, it's a native Google Sheets document
        return _get_google_sheets_all_start_dates(sheet)
    except Exception as e:
        # If gspread fails, try downloading as Excel file
        if "not supported for this document" in str(e):
            temp_path = None
            try:
                temp_path = _download_excel_file(sheet_link)
                return _get_excel_all_sheet_start_dates(temp_path)
            finally:
                # Clean up temporary file
                if temp_path and os.path.exists(temp_path):
                    os.unlink(temp_path)
        else:
            raise e


def _get_google_sheets_all_start_dates(sheet) -> Set[date]:

    start_dates: Set[date] = set()
    current_year = 2025

    target_sheets = []
    for ws in sheet.worksheets():
        title = (ws.title or "").strip()
        if title.isdigit() and len(title) == 4:
            target_sheets.append(title)
    if not target_sheets:
        target_sheets = [sheet.worksheets()[0].title]

    for worksheet_title in target_sheets:
        rows = sheet.worksheet(worksheet_title).get('A1:ZZ1000')
        header_row_idx = _find_first_header_row(rows)
        if header_row_idx is None:
            continue
        range_row_idx = header_row_idx + 1
        header_vals = rows[header_row_idx]
        range_vals = rows[range_row_idx] if range_row_idx < len(rows) else []
        month_spans = _collect_month_spans(header_vals)
        if not month_spans:
            continue
        for span in month_spans:
            mon = span["month"]
            for col_idx in range(span["start_col"], min(span["end_col"] + 1, len(range_vals))):
                token = (range_vals[col_idx] or "").strip()
                if not token or '-' not in token:
                    continue
                parsed = _parse_date_range(token, mon, current_year)
                if not parsed:
                    continue
                start_str, _ = parsed
                start_dates.add(datetime.strptime(start_str, "%Y/%m/%d").date())

    return start_dates
