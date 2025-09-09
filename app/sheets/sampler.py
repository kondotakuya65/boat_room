import os
import re
import csv
from typing import List, Tuple
import gspread

from .client import get_gspread_client

_SPREADSHEET_ID_RE = re.compile(r"/spreadsheets/d/([a-zA-Z0-9-_]+)")


def _ensure_dir(path: str):
    os.makedirs(path, exist_ok=True)


def extract_spreadsheet_id(sheet_link: str) -> str:
    m = _SPREADSHEET_ID_RE.search(sheet_link)
    if not m:
        raise ValueError("Invalid Google Sheets link")
    return m.group(1)


def _is_hidden(ws) -> bool:
    props = getattr(ws, "_properties", {}) or {}
    return bool(props.get("hidden", False))


def download_samples_for_spreadsheet(sheet_link: str, boat_name: str) -> List[str]:
    client = get_gspread_client()
    spreadsheet_id = extract_spreadsheet_id(sheet_link)
    sh = client.open_by_key(spreadsheet_id)
    saved: List[str] = []
    dest_root = os.path.join("data", "samples", boat_name)
    _ensure_dir(dest_root)
    for ws in sh.worksheets():
        if _is_hidden(ws):
            continue
        title = ws.title
        rows = ws.get_all_values()
        dest_path = os.path.join(dest_root, f"{title}.csv")
        with open(dest_path, "w", encoding="utf-8", newline="") as f:
            writer = csv.writer(f)
            for r in rows:
                writer.writerow(r)
        saved.append(dest_path)
    return saved


# Offline analyzer for Kanha group from a local XLSX sample
def analyze_kanha_xlsx(xlsx_path: str) -> dict:
    from openpyxl import load_workbook

    wb = load_workbook(filename=xlsx_path, data_only=True)
    ws = wb.active

    # convert to 0-based list of list strings
    rows: List[List[str]] = []
    for row in ws.iter_rows(values_only=True):
        rows.append([(str(c) if c is not None else "") for c in row])

    # find row with CABIN and ROOM in first two cells
    bands_row_idx = None
    for i, r in enumerate(rows):
        c0 = (r[0] if len(r) > 0 else '').strip().upper()
        c1 = (r[1] if len(r) > 1 else '').strip().upper()
        if c0.startswith('CABIN') and c1 == 'ROOM':
            bands_row_idx = i
            break

    result = {
        "bands_row_idx": bands_row_idx,
        "bands": [],
        "ot_columns": [],
    }

    if bands_row_idx is None:
        return result

    # detect bands using border thickness on that row
    def is_bold(cell) -> bool:
        b = getattr(cell, 'border', None)
        if not b:
            return False
        left = getattr(b, 'left', None)
        right = getattr(b, 'right', None)
        styles = {getattr(left, 'style', None), getattr(right, 'style', None)}
        return any(s in {"medium", "thick", "double"} for s in styles)

    # OT markers within band on same row
    ot_cols = []
    for j, cell in enumerate(ws[bands_row_idx + 1]):
        val = str(cell.value or '').strip().upper()
        if val == 'OT' or val == 'PRIVATE' or 'UPGRADE' in val:
            ot_cols.append(j)
    result["ot_columns"] = ot_cols

    bands: List[Tuple[int, int]] = []
    start = None
    row_cells = list(ws[bands_row_idx + 1])
    for j, cell in enumerate(row_cells):
        left_style = getattr(getattr(cell.border, 'left', None), 'style', None)
        if left_style in {"medium", "thick", "double"} and start is None:
            start = j
        right_style = getattr(getattr(cell.border, 'right', None), 'style', None)
        if start is not None and right_style in {"medium", "thick", "double"}:
            end = j
            # require at least one OT column inside
            if any(start <= c <= end for c in ot_cols):
                bands.append((start, end))
            start = None

    result["bands"] = bands
    return result
